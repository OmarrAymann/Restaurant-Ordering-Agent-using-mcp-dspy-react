import os
import smtplib
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from typing import List, Optional, Dict, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv
from mcp.server.fastmcp import FastMCP
from pydantic import BaseModel, Field

load_dotenv()

restaurant_service = FastMCP("RestaurantOrderingService")

SENDER_EMAIL = os.getenv("SENDER_EMAIL")
SENDER_PASSWORD = os.getenv("SENDER_PASSWORD")
DEFAULT_CHEF_EMAIL = os.getenv("CHEF_EMAIL", "chef@restaurant.com")

ORDERS_EXCEL_FILE = "restaurant_orders_log.xlsx"
TAX_RATE = 0.10


class MenuItemModel(BaseModel):
    """Represents a single item on the restaurant menu."""
    
    item_code: str = Field(description="Unique identifier for the menu item")
    dish_name: str = Field(description="Display name of the dish")
    category_type: str = Field(description="Menu category")
    detailed_description: str = Field(description="Detailed description of the dish")
    base_price: float = Field(description="Price in USD")
    ingredient_list: List[str] = Field(description="List of primary ingredients")
    dietary_labels: List[str] = Field(description="Dietary tags")
    is_available: bool = Field(description="Current availability status")
    prep_time_minutes: int = Field(description="Estimated preparation time")
    popularity_rating: int = Field(description="Popularity score (1-10)")


class OrderLineItem(BaseModel):
    """Represents a single item within an order."""
    
    item_code: str = Field(description="Menu item identifier")
    dish_name: str = Field(description="Name of the ordered dish")
    quantity_ordered: int = Field(default=1, description="Number of servings")
    special_instructions: Optional[str] = Field(default=None, description="Custom requests")
    unit_price: float = Field(description="Price per unit")


class CustomerInformation(BaseModel):
    """Customer contact and identification details."""
    
    full_name: str = Field(description="Customer's full name")
    table_or_room_number: int = Field(description="Table or room number")
    phone_number: str = Field(description="Contact phone number")
    email_address: Optional[str] = Field(default=None, description="Email for receipts")


class CompleteOrder(BaseModel):
    """Complete order structure with all details."""
    
    order_identifier: str = Field(description="Unique order ID")
    customer_info: CustomerInformation = Field(description="Customer details")
    line_items: List[OrderLineItem] = Field(description="Ordered items")
    subtotal_amount: float = Field(description="Pre-tax total")
    tax_amount: float = Field(description="Calculated tax")
    grand_total: float = Field(description="Final total including tax")
    additional_notes: Optional[str] = Field(default=None, description="Order-level notes")
    order_timestamp: str = Field(description="ISO format timestamp")
    order_status: str = Field(description="Current order status")


RESTAURANT_MENU_DATABASE: Dict[str, MenuItemModel] = {
    "APP_001": MenuItemModel(
        item_code="APP_001",
        dish_name="Mediterranean Bruschetta",
        category_type="appetizer",
        detailed_description="Artisan sourdough topped with heirloom tomatoes, fresh basil, and aged balsamic",
        base_price=9.99,
        ingredient_list=["sourdough bread", "heirloom tomatoes", "basil", "garlic", "balsamic vinegar", "olive oil"],
        dietary_labels=["vegetarian", "vegan-option"],
        is_available=True,
        prep_time_minutes=10,
        popularity_rating=8
    ),
    
    "APP_002": MenuItemModel(
        item_code="APP_002",
        dish_name="Crispy Buffalo Wings",
        category_type="appetizer",
        detailed_description="Double-fried chicken wings tossed in house buffalo sauce with celery and blue cheese dip",
        base_price=13.99,
        ingredient_list=["chicken wings", "buffalo sauce", "butter", "celery", "blue cheese"],
        dietary_labels=["spicy", "gluten-free-option"],
        is_available=True,
        prep_time_minutes=18,
        popularity_rating=9
    ),
    
    "APP_003": MenuItemModel(
        item_code="APP_003",
        dish_name="Spinach Artichoke Dip",
        category_type="appetizer",
        detailed_description="Creamy blend of spinach, artichokes, three cheeses, served with tortilla chips",
        base_price=11.49,
        ingredient_list=["spinach", "artichokes", "cream cheese", "mozzarella", "parmesan", "tortilla chips"],
        dietary_labels=["vegetarian"],
        is_available=True,
        prep_time_minutes=12,
        popularity_rating=8
    ),
    
    "MAIN_001": MenuItemModel(
        item_code="MAIN_001",
        dish_name="Pan-Seared Atlantic Salmon",
        category_type="main",
        detailed_description="Wild-caught salmon with lemon beurre blanc, roasted asparagus, and herb fingerling potatoes",
        base_price=26.99,
        ingredient_list=["atlantic salmon", "lemon", "butter", "white wine", "asparagus", "fingerling potatoes", "herbs"],
        dietary_labels=["gluten-free", "pescatarian"],
        is_available=True,
        prep_time_minutes=25,
        popularity_rating=10
    ),
    
    "MAIN_002": MenuItemModel(
        item_code="MAIN_002",
        dish_name="Classic Fettuccine Alfredo",
        category_type="main",
        detailed_description="Handmade fettuccine in rich parmesan cream sauce with grilled herb chicken breast",
        base_price=19.99,
        ingredient_list=["fettuccine pasta", "chicken breast", "heavy cream", "parmesan", "garlic", "butter"],
        dietary_labels=[],
        is_available=True,
        prep_time_minutes=22,
        popularity_rating=9
    ),
    
    "MAIN_003": MenuItemModel(
        item_code="MAIN_003",
        dish_name="Angus Ribeye Steak",
        category_type="main",
        detailed_description="12oz USDA Prime ribeye, garlic mashed potatoes, grilled broccolini, red wine reduction",
        base_price=34.99,
        ingredient_list=["ribeye steak", "potatoes", "butter", "broccolini", "red wine", "shallots"],
        dietary_labels=["gluten-free"],
        is_available=True,
        prep_time_minutes=30,
        popularity_rating=10
    ),
    
    "DESS_001": MenuItemModel(
        item_code="DESS_001",
        dish_name="Molten Chocolate Lava Cake",
        category_type="dessert",
        detailed_description="Warm Belgian chocolate cake with liquid center, vanilla bean ice cream, raspberry coulis",
        base_price=9.99,
        ingredient_list=["dark chocolate", "flour", "eggs", "butter", "vanilla ice cream", "raspberries"],
        dietary_labels=["vegetarian"],
        is_available=True,
        prep_time_minutes=14,
        popularity_rating=10
    ),
    
    "DESS_002": MenuItemModel(
        item_code="DESS_002",
        dish_name="New York Cheesecake",
        category_type="dessert",
        detailed_description="Classic creamy cheesecake with graham cracker crust, fresh berry compote",
        base_price=8.49,
        ingredient_list=["cream cheese", "graham crackers", "eggs", "sour cream", "mixed berries"],
        dietary_labels=["vegetarian"],
        is_available=True,
        prep_time_minutes=8,
        popularity_rating=9
    ),
    
    "DRINK_001": MenuItemModel(
        item_code="DRINK_001",
        dish_name="Fresh-Squeezed Lemonade",
        category_type="drink",
        detailed_description="House-made lemonade with organic lemons, cane sugar, and fresh mint",
        base_price=4.99,
        ingredient_list=["lemons", "cane sugar", "water", "fresh mint"],
        dietary_labels=["vegan", "vegetarian", "gluten-free"],
        is_available=True,
        prep_time_minutes=5,
        popularity_rating=8
    ),
    
    "DRINK_002": MenuItemModel(
        item_code="DRINK_002",
        dish_name="Craft Root Beer Float",
        category_type="drink",
        detailed_description="Artisan root beer with premium vanilla ice cream",
        base_price=6.49,
        ingredient_list=["root beer", "vanilla ice cream"],
        dietary_labels=["vegetarian"],
        is_available=True,
        prep_time_minutes=3,
        popularity_rating=7
    ),
}

active_orders_database: Dict[str, Dict[str, Any]] = {}


@restaurant_service.tool()
def fetch_menu(category_filter: str = "all") -> List[MenuItemModel] | str:
    """
    Retrieve menu items by category.
    
    Args:
        category_filter: Category to filter by (appetizer, main, dessert, drink, all)
        
    Returns:
        List of menu items or error message
    """
    normalized_category = category_filter.lower().strip()
    
    if normalized_category == "all":
        return list(RESTAURANT_MENU_DATABASE.values())
    
    filtered_items = [
        item for item in RESTAURANT_MENU_DATABASE.values()
        if item.category_type.lower() == normalized_category
    ]
    
    if not filtered_items:
        return f"No menu items found in category: {category_filter}. Available categories: appetizer, main, dessert, drink, all"
    
    return filtered_items


@restaurant_service.tool()
def calculate_order_total(item_codes: List[str]) -> Dict[str, float]:
    """
    Calculate the total cost for a list of menu items including tax.
    
    Args:
        item_codes: List of menu item codes
        
    Returns:
        Dictionary with subtotal, tax, and grand total
    """
    subtotal = 0.0
    
    for item_code in item_codes:
        normalized_code = item_code.upper().strip()
        
        if normalized_code not in RESTAURANT_MENU_DATABASE:
            raise ValueError(f"Menu item '{item_code}' not found. Please use valid item codes.")
        
        subtotal += RESTAURANT_MENU_DATABASE[normalized_code].base_price
    
    tax = subtotal * TAX_RATE
    grand_total = subtotal + tax
    
    return {
        "subtotal": round(subtotal, 2),
        "tax": round(tax, 2),
        "grand_total": round(grand_total, 2)
    }


@restaurant_service.tool()
def create_new_order(
    customer_name: str,
    service_location: str,
    contact_phone: str,
    item_codes: List[str],
    customer_email: Optional[str] = None
) -> Dict[str, Any]:
    """
    Create a new order and save it to the system.
    
    Args:
        customer_name: Full name of the customer
        service_location: Table number or delivery location
        contact_phone: Customer's phone number
        item_codes: List of menu item codes to order
        customer_email: Optional email for receipt
        
    Returns:
        Dictionary with success status, order ID, and order details
    """
    order_number = len(active_orders_database) + 1
    order_id = f"ORD-{order_number:05d}"
    
    for item_code in item_codes:
        normalized_code = item_code.upper().strip()
        if normalized_code not in RESTAURANT_MENU_DATABASE:
            return {
                "success": False,
                "error": f"Invalid menu item: '{item_code}'. Please check the menu."
            }
    
    pricing_data = calculate_order_total(item_codes)
    
    new_order = {
        "order_id": order_id,
        "customer_name": customer_name,
        "service_location": service_location,
        "contact_phone": contact_phone,
        "customer_email": customer_email,
        "ordered_items": item_codes,
        "subtotal": pricing_data["subtotal"],
        "tax": pricing_data["tax"],
        "grand_total": pricing_data["grand_total"],
        "timestamp": datetime.now().isoformat(),
        "status": "pending"
    }
    
    active_orders_database[order_id] = new_order
    
    return {
        "success": True,
        "order_id": order_id,
        "message": f"Order successfully created for {customer_name} at {service_location}",
        "order_details": new_order
    }


@restaurant_service.tool()
def send_order_to_kitchen(
    order_id: str,
    chef_email: str = DEFAULT_CHEF_EMAIL
) -> Dict[str, Any]:
    """
    Send order details to the kitchen via email notification.
    
    Args:
        order_id: The unique order identifier
        chef_email: Email address of the kitchen/chef
        
    Returns:
        Dictionary with success status and details
    """
    if order_id not in active_orders_database:
        return {
            "success": False,
            "error": f"Order {order_id} not found in system"
        }
    
    order_data = active_orders_database[order_id]
    
    items_text = "\n".join([
        f"  • {RESTAURANT_MENU_DATABASE[item_code].dish_name} (Code: {item_code})"
        for item_code in order_data["ordered_items"]
    ])
    
    email_subject = f"🔔 NEW ORDER - {order_id} | {order_data['service_location']}"
    
    email_body = f"""
╔════════════════════════════════════════════════════════╗
║               NEW ORDER NOTIFICATION                    ║
╚════════════════════════════════════════════════════════╝

📋 ORDER DETAILS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Order ID:        {order_id}
Timestamp:       {order_data['timestamp']}

👤 CUSTOMER INFORMATION
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Name:            {order_data['customer_name']}
Location:        {order_data['service_location']}
Phone:           {order_data['contact_phone']}

🍽️ ORDERED ITEMS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{items_text}

💰 PRICING
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Subtotal:        ${order_data['subtotal']:.2f}
Tax (10%):       ${order_data['tax']:.2f}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
GRAND TOTAL:     ${order_data['grand_total']:.2f}

⚡ Please prepare this order with priority.
    """
    
    try:
        message = MIMEMultipart()
        message['From'] = SENDER_EMAIL
        message['To'] = chef_email
        message['Subject'] = email_subject
        message.attach(MIMEText(email_body, 'plain'))
        
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp_server:
            smtp_server.login(SENDER_EMAIL, SENDER_PASSWORD)
            smtp_server.send_message(message)
        
        active_orders_database[order_id]["status"] = "sent_to_kitchen"
        
        return {
            "success": True,
            "message": f"Order {order_id} successfully sent to kitchen",
            "email_recipient": chef_email,
            "order_summary": email_body
        }
    
    except Exception as error:
        return {
            "success": False,
            "error": f"Failed to send email: {str(error)}",
            "order_summary": email_body,
            "note": "Order saved but email notification failed."
        }


@restaurant_service.tool()
def save_order_to_excel(
    order_id: str,
    customer_name: str,
    contact_phone: str,
    service_location: str,
    item_codes: List[str]
) -> Dict[str, Any]:
    """
    Save order details to an Excel file for record-keeping.
    
    Args:
        order_id: Unique order identifier
        customer_name: Customer's full name
        contact_phone: Phone number
        service_location: Table/room number
        item_codes: List of ordered item codes
        
    Returns:
        Dictionary with success status and file path
    """
    try:
        if os.path.exists(ORDERS_EXCEL_FILE):
            workbook = openpyxl.load_workbook(ORDERS_EXCEL_FILE)
            worksheet = workbook.active
        else:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Orders Log"
            
            headers = ["Order ID", "Timestamp", "Customer Name", "Phone Number", 
                       "Location", "Items Ordered", "Total Amount"]
            worksheet.append(headers)
            
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
        
        items_display = ", ".join([
            RESTAURANT_MENU_DATABASE[code].dish_name 
            for code in item_codes
        ])
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        pricing = calculate_order_total(item_codes)
        
        new_row = [
            order_id,
            timestamp,
            customer_name,
            contact_phone,
            service_location,
            items_display,
            f"${pricing['grand_total']:.2f}"
        ]
        worksheet.append(new_row)
        
        column_widths = [15, 20, 25, 18, 15, 50, 15]
        for idx, width in enumerate(column_widths, start=1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
        
        workbook.save(ORDERS_EXCEL_FILE)
        
        return {
            "success": True,
            "message": f"Order {order_id} saved to Excel log",
            "file_path": ORDERS_EXCEL_FILE
        }
    
    except Exception as error:
        return {
            "success": False,
            "error": f"Failed to save to Excel: {str(error)}"
        }


@restaurant_service.tool()
def get_order_status(order_id: str) -> Dict[str, Any]:
    """
    Retrieve the current status and details of an order.
    
    Args:
        order_id: The unique order identifier
        
    Returns:
        Order details or error message
    """
    if order_id not in active_orders_database:
        return {
            "success": False,
            "error": f"Order {order_id} not found"
        }
    
    return {
        "success": True,
        "order": active_orders_database[order_id]
    }


if __name__ == "__main__":
    print("🚀 Starting Enhanced Restaurant MCP Server...")
    restaurant_service.run()