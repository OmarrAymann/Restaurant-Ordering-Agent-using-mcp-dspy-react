import os
import smtplib
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from typing import List, Optional, Dict, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv
from mcp.server.fastmcp import FastMCP
from pydantic import BaseModel, Field

load_dotenv()

restaurant_service = FastMCP("RestaurantOrderingService")
SENDER_EMAIL = os.getenv("SENDER_EMAIL")
SENDER_PASSWORD = os.getenv("SENDER_PASSWORD")
restaurant_service = os.getenv("RES_EMAIL", "chef@restaurant.com")
ORDERS_EXCEL_FILE = "restaurant_orders_log.xlsx"
TAX_RATE = 0.14


class MenuItemModel(BaseModel):
    """Represents a single item on the restaurant menu."""
    
    item_code: str
    dish_name: str
    category_type: str
    detailed_description: str
    base_price: float
    ingredient_list: List[str]
    dietary_labels: List[str]
    is_available: bool
    prep_time_minutes: int
    popularity_rating: int


RESTAURANT_MENU_DATABASE: Dict[str, MenuItemModel] = {
    "APP_001": MenuItemModel(
        item_code="APP_001",
        dish_name="Falafel",
        category_type="appetizer",
        detailed_description="Crispy chickpea balls served with tahini sauce and fresh herbs",
        base_price=9.99,
        ingredient_list=["chickpeas", "parsley", "garlic", "onion", "tahini", "spices"],
        dietary_labels=["vegetarian", "vegan"],
        is_available=True,
        prep_time_minutes=10,
        popularity_rating=9
    ),
    
    "APP_002": MenuItemModel(
        item_code="APP_002",
        dish_name="Koshari Balls",
        category_type="appetizer",
        detailed_description="Mini koshari bites with rice, lentils, pasta, and tomato sauce",
        base_price=13.99,
        ingredient_list=["rice", "lentils", "pasta", "tomato sauce", "fried onions"],
        dietary_labels=["vegetarian", "vegan"],
        is_available=True,
        prep_time_minutes=15,
        popularity_rating=8
    ),
    
    "APP_003": MenuItemModel(
        item_code="APP_003",
        dish_name="Sambousek",
        category_type="appetizer",
        detailed_description="Savory pastry filled with cheese or minced meat, lightly fried",
        base_price=11.49,
        ingredient_list=["flour", "cheese", "minced meat", "onion", "spices"],
        dietary_labels=["vegetarian-option"],
        is_available=True,
        prep_time_minutes=12,
        popularity_rating=8
    ),
    
    "MAIN_001": MenuItemModel(
        item_code="MAIN_001",
        dish_name="Grilled Kofta",
        category_type="main",
        detailed_description="Seasoned beef and lamb skewers served with rice and vegetables",
        base_price=26.99,
        ingredient_list=["beef", "lamb", "onion", "parsley", "spices", "rice", "vegetables"],
        dietary_labels=["gluten-free-option"],
        is_available=True,
        prep_time_minutes=25,
        popularity_rating=10
    ),
    
    "MAIN_002": MenuItemModel(
        item_code="MAIN_002",
        dish_name="Molokhia with Rabbit",
        category_type="main",
        detailed_description="Traditional Egyptian molokhia stew served with rice and tender rabbit meat",
        base_price=19.99,
        ingredient_list=["molokhia leaves", "rabbit", "garlic", "coriander", "rice", "spices"],
        dietary_labels=[],
        is_available=True,
        prep_time_minutes=22,
        popularity_rating=9
    ),
    
    "MAIN_003": MenuItemModel(
        item_code="MAIN_003",
        dish_name="Grilled Hamam (Pigeon)",
        category_type="main",
        detailed_description="Stuffed pigeon grilled with Egyptian spices, served with rice",
        base_price=34.99,
        ingredient_list=["pigeon", "spices", "rice", "onion", "garlic", "herbs"],
        dietary_labels=["gluten-free-option"],
        is_available=True,
        prep_time_minutes=30,
        popularity_rating=10
    ),
    
    "DESS_001": MenuItemModel(
        item_code="DESS_001",
        dish_name="Basbousa",
        category_type="dessert",
        detailed_description="Sweet semolina cake topped with almonds and soaked in syrup",
        base_price=9.99,
        ingredient_list=["semolina", "sugar", "almonds", "butter", "milk", "syrup"],
        dietary_labels=["vegetarian"],
        is_available=True,
        prep_time_minutes=14,
        popularity_rating=10
    ),
    
    "DESS_002": MenuItemModel(
        item_code="DESS_002",
        dish_name="Konafa",
        category_type="dessert",
        detailed_description="Shredded phyllo pastry layered with sweet cheese or cream, soaked in syrup",
        base_price=8.49,
        ingredient_list=["kataifi pastry", "cheese", "cream", "sugar", "syrup"],
        dietary_labels=["vegetarian"],
        is_available=True,
        prep_time_minutes=12,
        popularity_rating=9
    ),
    
    "DRINK_001": MenuItemModel(
        item_code="DRINK_001",
        dish_name="Karkadeh",
        category_type="drink",
        detailed_description="Refreshing hibiscus tea served hot or cold",
        base_price=4.99,
        ingredient_list=["hibiscus petals", "sugar", "water"],
        dietary_labels=["vegan", "vegetarian", "gluten-free"],
        is_available=True,
        prep_time_minutes=5,
        popularity_rating=8
    ),
    
    "DRINK_002": MenuItemModel(
        item_code="DRINK_002",
        dish_name="Sugarcane Juice",
        category_type="drink",
        detailed_description="Freshly pressed sugarcane juice served chilled",
        base_price=15,
        ingredient_list=["sugarcane", "ice", "lemon"],
        dietary_labels=["vegan", "vegetarian", "gluten-free"],
        is_available=True,
        prep_time_minutes=3,
        popularity_rating=9
    ),
}


active_orders_database: Dict[str, Dict[str, Any]] = {}


@restaurant_service.tool()
def fetch_menu(category_filter: str = "all"):
    """Retrieve menu items by category."""
    normalized_category = category_filter.lower().strip()
    
    if normalized_category == "all":
        return list(RESTAURANT_MENU_DATABASE.values())
    
    filtered_items = [
        item for item in RESTAURANT_MENU_DATABASE.values()
        if item.category_type.lower() == normalized_category
    ]
    
    if not filtered_items:
        return f"No menu items found in category: {category_filter}"
    
    return filtered_items


@restaurant_service.tool()
def calculate_order_total(item_codes: List[str]) -> Dict[str, float]:
    """Calculate the total cost for a list of menu items including tax."""
    subtotal = 0.0
    
    for item_code in item_codes:
        normalized_code = item_code.upper().strip()
        
        if normalized_code not in RESTAURANT_MENU_DATABASE:
            raise ValueError(f"Menu item '{item_code}' not found.")
        
        subtotal += RESTAURANT_MENU_DATABASE[normalized_code].base_price
    
    tax = subtotal * TAX_RATE
    grand_total = subtotal + tax
    
    return {
        "subtotal": round(subtotal, 2),
        "tax": round(tax, 2),
        "grand_total": round(grand_total, 2)
    }


@restaurant_service.tool()
def create_new_order(
    customer_name: str,
    service_location: str,
    contact_phone: str,
    item_codes: List[str],
    customer_email: Optional[str] = None
) -> Dict[str, Any]:
    """Create a new order and save it to the system."""
    order_number = len(active_orders_database) + 1
    order_id = f"ORD-{order_number:05d}"
    
    for item_code in item_codes:
        normalized_code = item_code.upper().strip()
        if normalized_code not in RESTAURANT_MENU_DATABASE:
            return {"success": False, "error": f"Invalid menu item: '{item_code}'"}
    
    pricing_data = calculate_order_total(item_codes)
    
    new_order = {
        "order_id": order_id,
        "customer_name": customer_name,
        "service_location": service_location,
        "contact_phone": contact_phone,
        "customer_email": customer_email,
        "ordered_items": item_codes,
        "subtotal": pricing_data["subtotal"],
        "tax": pricing_data["tax"],
        "grand_total": pricing_data["grand_total"],
        "timestamp": datetime.now().isoformat(),
        "status": "pending"
    }
    
    active_orders_database[order_id] = new_order
    
    return {
        "success": True,
        "order_id": order_id,
        "message": f"Order created for {customer_name} at {service_location}",
        "order_details": new_order
    }


@restaurant_service.tool()
def send_order_to_kitchen(order_id: str, chef_email: str = DEFAULT_CHEF_EMAIL) -> Dict[str, Any]:
    """Send order details to the kitchen via email."""
    if order_id not in active_orders_database:
        return {"success": False, "error": f"Order {order_id} not found"}
    
    order_data = active_orders_database[order_id]
    
    items_text = "\n".join([
        f"  • {RESTAURANT_MENU_DATABASE[item_code].dish_name} (Code: {item_code})"
        for item_code in order_data["ordered_items"]
    ])
    
    email_subject = f"🔔 NEW ORDER - {order_id} | {order_data['service_location']}"
    
    email_body = f"""
╔════════════════════════════════════════════════════════╗
║               NEW ORDER NOTIFICATION                    ║
╚════════════════════════════════════════════════════════╝

📋 ORDER DETAILS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Order ID:        {order_id}
Timestamp:       {order_data['timestamp']}

👤 CUSTOMER INFORMATION
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Name:            {order_data['customer_name']}
Location:        {order_data['service_location']}
Phone:           {order_data['contact_phone']}

🍽️ ORDERED ITEMS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{items_text}

💰 PRICING
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Subtotal:        ${order_data['subtotal']:.2f}
Tax (10%):       ${order_data['tax']:.2f}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
GRAND TOTAL:     ${order_data['grand_total']:.2f}
    """
    
    try:
        message = MIMEMultipart()
        message['From'] = SENDER_EMAIL
        message['To'] = chef_email
        message['Subject'] = email_subject
        message.attach(MIMEText(email_body, 'plain'))
        
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp_server:
            smtp_server.login(SENDER_EMAIL, SENDER_PASSWORD)
            smtp_server.send_message(message)
        
        active_orders_database[order_id]["status"] = "sent_to_kitchen"
        
        return {
            "success": True,
            "message": f"Order {order_id} sent to kitchen",
            "email_recipient": chef_email
        }
    
    except Exception as error:
        return {
            "success": False,
            "error": f"Failed to send email: {str(error)}",
            "note": "Order saved but email failed"
        }


@restaurant_service.tool()
def save_order_to_excel(
    order_id: str,
    customer_name: str,
    contact_phone: str,
    service_location: str,
    item_codes: List[str]
) -> Dict[str, Any]:
    """Save order details to Excel file."""
    try:
        if os.path.exists(ORDERS_EXCEL_FILE):
            workbook = openpyxl.load_workbook(ORDERS_EXCEL_FILE)
            worksheet = workbook.active
        else:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Orders Log"
            
            headers = ["Order ID", "Timestamp", "Customer Name", "Phone Number", 
                    "Location", "Items Ordered", "Total Amount"]
            worksheet.append(headers)
            
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
        
        items_display = ", ".join([
            RESTAURANT_MENU_DATABASE[code].dish_name 
            for code in item_codes
        ])
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        pricing = calculate_order_total(item_codes)
        
        new_row = [
            order_id,
            timestamp,
            customer_name,
            contact_phone,
            service_location,
            items_display,
            f"${pricing['grand_total']:.2f}"
        ]
        worksheet.append(new_row)
        
        workbook.save(ORDERS_EXCEL_FILE)
        
        return {
            "success": True,
            "message": f"Order {order_id} saved to Excel",
            "file_path": ORDERS_EXCEL_FILE
        }
    
    except Exception as error:
        return {
            "success": False,
            "error": f"Failed to save to Excel: {str(error)}"
        }


@restaurant_service.tool()
def get_order_status(order_id: str) -> Dict[str, Any]:
    """Retrieve order status and details."""
    if order_id not in active_orders_database:
        return {"success": False, "error": f"Order {order_id} not found"}
    
    return {
        "success": True,
        "order": active_orders_database[order_id]
    }


if __name__ == "__main__":
    print("Starting Restaurant MCP Server...")
    restaurant_service.run()